#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


CELL_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def norm_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().upper())


def norm_part_number(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.strip().upper().replace("O", "0"))


def to_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_shared_strings(zf: ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    out: List[str] = []
    for si in root.findall(f"{CELL_NS}si"):
        parts = [t.text or "" for t in si.findall(f".//{CELL_NS}t")]
        out.append("".join(parts))
    return out


def parse_sheet_cells(zf: ZipFile, sheet_xml_path: str, shared_strings: List[str]) -> List[List[str]]:
    root = ET.fromstring(zf.read(sheet_xml_path))
    rows = []
    for row in root.findall(f".//{CELL_NS}sheetData/{CELL_NS}row"):
        data = {}
        max_col = 0
        for c in row.findall(f"{CELL_NS}c"):
            ref = c.attrib.get("r", "")
            col = 0
            for ch in ref:
                if "A" <= ch <= "Z":
                    col = col * 26 + (ord(ch) - ord("A") + 1)
                else:
                    break
            max_col = max(max_col, col)
            cell_type = c.attrib.get("t")
            v = c.find(f"{CELL_NS}v")
            text = ""
            if v is not None and v.text is not None:
                if cell_type == "s":
                    idx = int(v.text)
                    text = shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
                else:
                    text = v.text
            is_text = c.find(f"{CELL_NS}is/{CELL_NS}t")
            if is_text is not None and is_text.text:
                text = is_text.text
            if col > 0:
                data[col] = text
        if max_col > 0:
            rows.append([data.get(i, "") for i in range(1, max_col + 1)])
    return rows


def extract_part_numbers_from_part_list(part_list_xlsx: Path) -> List[str]:
    with ZipFile(part_list_xlsx, "r") as zf:
        shared = parse_shared_strings(zf)
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
        }

        numbers: List[str] = []
        for idx, sheet in enumerate(workbook.findall(f".//{CELL_NS}sheets/{CELL_NS}sheet")):
            if idx >= 12:
                break
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if rid not in rel_map:
                continue
            target = rel_map[rid]
            sheet_path = f"xl/{target}" if not target.startswith("xl/") else target
            rows = parse_sheet_cells(zf, sheet_path, shared)
            for row in rows[:600]:
                for val in row[:20]:
                    text = str(val).strip()
                    if re.search(r"\d{4,}[-/][A-Z0-9]+", text.upper()):
                        numbers.append(text)
        return sorted(set(numbers))


@dataclass
class DefectObservation:
    line: str
    part_number: str
    defect_name: str
    qty: int
    sheet_name: str


def parse_daily_ng_workbook(path: Path, line: str) -> List[DefectObservation]:
    wb = load_workbook(path, data_only=True, read_only=True)
    observations: List[DefectObservation] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 900), max_col=20, values_only=True))
        legend: Dict[str, str] = {}
        for idx, row in enumerate(rows[:100], start=1):
            a1 = row[2] if len(row) > 2 else None
            b1 = row[3] if len(row) > 3 else None
            a2 = row[5] if len(row) > 5 else None
            b2 = row[6] if len(row) > 6 else None
            if isinstance(a1, str) and len(a1.strip()) == 1 and b1 is not None:
                key = a1.strip().upper()
                val = str(b1).strip()
                if val and "TOTAL" not in val.upper():
                    legend[key] = val
            if isinstance(a2, str) and len(a2.strip()) == 1 and b2 is not None:
                key = a2.strip().upper()
                val = str(b2).strip()
                if val and "TOTAL" not in val.upper():
                    legend[key] = val

        current_part_number: Optional[str] = None
        for idx, row in enumerate(rows, start=1):
            no_cell = row[1] if len(row) > 1 else None
            item_cell = row[2] if len(row) > 2 else None
            code_cell = row[7] if len(row) > 7 else None

            if isinstance(no_cell, (int, float)) and item_cell is not None:
                current_part_number = None

            if isinstance(item_cell, str):
                item_text = item_cell.strip().upper()
                if re.search(r"\d{4,}[-/][A-Z0-9]+", item_text):
                    current_part_number = item_cell.strip()

            if current_part_number is None:
                continue

            if not isinstance(code_cell, str):
                continue
            code = code_cell.strip().upper()
            if len(code) != 1 or code not in legend:
                continue

            qty = sum(to_int(row[c]) for c in range(8, min(len(row), 17)))
            if qty <= 0:
                continue

            observations.append(
                DefectObservation(
                    line=line,
                    part_number=current_part_number,
                    defect_name=legend[code],
                    qty=qty,
                    sheet_name=ws.title,
                ),
            )
    return observations


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--press", required=True)
    parser.add_argument("--sewing", required=True)
    parser.add_argument("--part-list", required=True)
    parser.add_argument("--requirement", required=True)
    parser.add_argument("--mapping", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--skip-part-list", action="store_true")
    args = parser.parse_args()

    press_path = Path(args.press)
    sewing_path = Path(args.sewing)
    part_list_path = Path(args.part_list)
    requirement_path = Path(args.requirement)
    mapping_path = Path(args.mapping)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    press_obs = parse_daily_ng_workbook(press_path, "press")
    sewing_obs = parse_daily_ng_workbook(sewing_path, "sewing")
    observations = press_obs + sewing_obs

    with mapping_path.open("r", encoding="utf-8") as fp:
        mapping = json.load(fp)

    part_by_norm = {}
    part_number_set = set()
    for p in mapping.get("parts", []):
        part_no = str(p.get("part_number", "")).strip()
        if not part_no:
            continue
        norm = norm_part_number(part_no)
        part_by_norm[norm] = p
        part_number_set.add(part_no)

    part_list_numbers: List[str] = []
    part_list_warning = None
    if not args.skip_part_list:
        try:
            part_list_numbers = extract_part_numbers_from_part_list(part_list_path)
        except Exception as exc:  # noqa: BLE001
            part_list_warning = f"failed to parse part list: {type(exc).__name__}: {exc}"
    else:
        part_list_warning = "part list parsing skipped by flag"

    req_wb = load_workbook(requirement_path, data_only=True, read_only=True)
    req_numbers = set()
    req_ws = req_wb.worksheets[0]
    for r in range(1, req_ws.max_row + 1):
        for c in range(1, min(req_ws.max_column, 12) + 1):
            v = req_ws.cell(r, c).value
            if isinstance(v, str) and re.search(r"\d{4,}[-/][A-Z0-9]+", v.strip().upper()):
                req_numbers.add(v.strip())

    stats = defaultdict(lambda: {"qty": 0, "days": set()})
    unmatched_parts = set()
    for ob in observations:
        key = (ob.line, ob.part_number.strip(), ob.defect_name.strip())
        stats[key]["qty"] += ob.qty
        stats[key]["days"].add(ob.sheet_name)
        if norm_part_number(ob.part_number) not in part_by_norm:
            unmatched_parts.add(ob.part_number)

    material_risk = defaultdict(lambda: {"qty": 0, "parts": set(), "samples": 0})
    for (line, part_number, defect_name), data in stats.items():
        part = part_by_norm.get(norm_part_number(part_number))
        if not part:
            continue
        materials = part.get("materials", [])
        for mat in materials:
            mat_name = str(mat.get("material_name", "")).strip()
            if not mat_name:
                continue
            mk = (line, mat_name, defect_name.strip())
            material_risk[mk]["qty"] += data["qty"]
            material_risk[mk]["parts"].add(part.get("uniq_no"))
            material_risk[mk]["samples"] += 1

    part_item_defect_stats = []
    for (line, part_number, defect_name), data in sorted(stats.items()):
        part_item_defect_stats.append(
            {
                "line": line,
                "part_number": part_number,
                "part_number_norm": norm_part_number(part_number),
                "defect_name": defect_name,
                "defect_name_norm": norm_text(defect_name),
                "occurrence_qty": int(data["qty"]),
                "affected_days": len(data["days"]),
                "last_seen_sheet": max(data["days"]) if data["days"] else None,
            },
        )

    material_item_defect_risk = []
    for (line, material_name, defect_name), data in sorted(material_risk.items()):
        sample_size = max(1, data["samples"])
        risk_score = round(float(data["qty"]) / float(sample_size), 4)
        material_item_defect_risk.append(
            {
                "line": line,
                "material_name": material_name,
                "material_name_norm": norm_text(material_name),
                "defect_name": defect_name,
                "defect_name_norm": norm_text(defect_name),
                "risk_score": risk_score,
                "affected_parts": len(data["parts"]),
                "sample_size": sample_size,
            },
        )

    payload = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "source_files": {
            "daily_ng_press": str(press_path),
            "daily_ng_sewing": str(sewing_path),
            "part_list": str(part_list_path),
            "part_requirement_list": str(requirement_path),
            "mapping_json": str(mapping_path),
        },
        "summary": {
            "total_observations": len(observations),
            "distinct_part_defect_pairs": len(part_item_defect_stats),
            "distinct_material_defect_risk": len(material_item_defect_risk),
            "unmatched_parts_count": len(unmatched_parts),
            "part_list_numbers_count": len(part_list_numbers),
            "requirement_numbers_count": len(req_numbers),
            "mapping_part_numbers_count": len(part_number_set),
        },
        "warnings": [w for w in [part_list_warning] if w],
        "unmatched_part_numbers_from_daily_ng": sorted(unmatched_parts),
        "part_item_defect_stats": part_item_defect_stats,
        "material_item_defect_risk": material_item_defect_risk,
    }

    with output_path.open("w", encoding="utf-8") as fp:
        json.dump(payload, fp, ensure_ascii=False, indent=2)

    print(f"written: {output_path}")
    print(
        "summary:",
        payload["summary"],
    )


if __name__ == "__main__":
    main()
